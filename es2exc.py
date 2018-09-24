import argparse
import json
import logging
import sys
import yaml
import os
from elasticsearch import Elasticsearch
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference, BarChart
from openpyxl.chart.series import DataPoint
from openpyxl.styles import Font, colors
from openpyxl.utils import get_column_letter

__author__ = 'Miso Mijatovic'
__version__ = '1.0'

e_logger = logging.Logger(__name__)
e_handler = logging.FileHandler(os.path.join(os.path.dirname(os.path.realpath(__file__)), 'es2exc.log'))
e_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
e_handler.setFormatter(e_formatter)
e_logger.addHandler(e_handler)


def swap_columns(w, col1, col2):
    temp = []

    max_length = len(w[get_column_letter(col1)]) + 1
    for _ in range(1, max_length):
        col = w['{}{}'.format(get_column_letter(col1), _)]
        temp.append(col.value)
        col.value = ''

    max_length = len(w[get_column_letter(col2)]) + 1
    for _ in range(1, max_length):
        w['{}{}'.format(get_column_letter(col1), _)].value = w['{}{}'.format(get_column_letter(col2), _)].value

    last = 1
    max_length = len(w[get_column_letter(col2)]) + 1
    for _ in range(1, max_length):
        w['{}{}'.format(get_column_letter(col2), _)].value = temp[_ - 1]
        last = _
    last += 1
    if last < max_length:
        for i in range(last, max_length):
            w['{}{}'.format(get_column_letter(col2), i)].value = ''


def order_columns(w_sheet, c_values):
    """
    Order columns in alphabetical order by the header
    """
    columns_current = [str(_.value) for _ in w_sheet['1']]
    columns_ordered = sorted(c_values.keys())
    index = 0
    while index < len(columns_ordered):
        if columns_current == columns_ordered:
            break
        header_this = columns_current[index]
        header_ordered = columns_ordered[index]
        if header_this != header_ordered:
            index_ordered = columns_current.index(header_ordered)
            swap_columns(w_sheet, index + 1, index_ordered + 1)
            c_values[header_this] = index_ordered + 1, c_values[header_this][1]
            c_values[header_ordered] = index + 1, c_values[header_ordered][1]
            columns_current = [str(_.value) for _ in w_sheet['1']]
        else:
            index += 1
    return w_sheet, c_values


def new_key(upper, new):
    if upper != '':
        return upper + '.' + str(new)
    else:
        return new


def loop_on_nested_dict(the_element, upper_key=''):
    if isinstance(the_element, dict):
        for key, value in the_element.items():
            for _ in loop_on_nested_dict(value, new_key(upper_key, key)):
                yield _
    elif isinstance(the_element, list):
        for i in range(0, len(the_element)):
            for _ in loop_on_nested_dict(the_element[i], '{}[{}]'.format(upper_key, i)):
                yield _
    else:
        yield upper_key, the_element


def data_from_aggs(es_buckets):
    data = []
    width = 8
    for bucket in es_buckets:
        data.append([bucket['key'], bucket['doc_count']])
        if 60 > len(bucket['key']) > width:
            width = len(bucket['key'])
    return data, width


def get_next_sheet(wrkb, name):
    """
    Create a new sheet
    If only the first sheet 'Sheet' is present, rename it
    """
    if u'Sheet' in wrkb.sheetnames:
        this_sheet = wrkb.get_sheet_by_name(u'Sheet')
        this_sheet.title = name
    else:
        wrkb.create_sheet(name)


def main():

    es_response = es_client.search(index=args['index'], body=query)
    e_logger.info('got {} hits from es!'.format(len(es_response['hits']['hits'])))

    wb = Workbook()
    header_font = Font(color=colors.BLUE, bold=True)

    if es_response['hits']['hits']:

        get_next_sheet(wb, 'hits')
        ws = wb.get_sheet_by_name('hits')
        column_next = 1
        column_values = {}  # {'column_key':(column_index, column_width)}
        row_next = 2

        # add total hits sheet
        for hit in es_response['hits']['hits']:
            for k, v in loop_on_nested_dict(hit['_source']):
                if k not in column_values:
                    column_values[k] = (column_next, len(k))
                    column_next += 1
                    c = ws.cell(row=1, column=column_values[k][0], value=k)
                    c.font = header_font
                ws.cell(row=row_next, column=column_values[k][0], value=v)
                if column_values[k][1] < len(v) < 60:
                    column_values[k] = column_values[k][0], len(v)
            row_next += 1

        ws, column_values = order_columns(ws, column_values)

        for column_index, column_width in column_values.values():
            ws.column_dimensions[get_column_letter(column_index)].width = column_width

    # add a sheet for every aggregation
    if 'aggregations' in es_response.keys():
        for agg_name in es_response['aggregations'].keys():
            column_width_header = len(agg_name)
            get_next_sheet(wb, agg_name)
            ws = wb.get_sheet_by_name(agg_name)
            row = 1
            c = ws.cell(row=row, column=1, value=agg_name)
            c.font = header_font
            c = ws.cell(row=row, column=2, value='count')
            c.font = header_font
            aggs_data, column_width = data_from_aggs(es_response['aggregations'][agg_name]['buckets'])
            for row in aggs_data:
                ws.append(row)
            if column_width_header > column_width:
                ws.column_dimensions['A'].width = column_width_header
            else:
                ws.column_dimensions['A'].width = column_width

            # add pie chart
            if args['piechart']:
                pie_chart = PieChart()
                labels = Reference(ws, min_col=1, min_row=2, max_row=len(aggs_data) + 1)
                chart_data = Reference(ws, min_col=2, min_row=1, max_row=len(aggs_data) + 1)
                pie_chart.add_data(chart_data, titles_from_data=True)
                pie_chart.set_categories(labels)
                pie_chart.title = agg_name

                # Cut the first slice out of the pie
                pie_slice = DataPoint(idx=0, explosion=20)
                pie_chart.series[0].data_points = [pie_slice]

                ws.add_chart(pie_chart, "D2")

            # add bar chart
            if args['barchart']:
                bar_chart = BarChart()
                bar_chart.type = "col"
                bar_chart.style = 10
                bar_chart.title = agg_name
                bar_chart.y_axis.title = 'count'

                chart_data = Reference(ws, min_col=2, min_row=1, max_row=len(aggs_data) + 1)
                cats = Reference(ws, min_col=1, min_row=2, max_row=len(aggs_data) + 1)
                bar_chart.add_data(chart_data, titles_from_data=True)
                bar_chart.set_categories(cats)
                bar_chart.shape = 4
                ws.add_chart(bar_chart, "D20")

    wb.save(args['output'])
    e_logger.info('saved file {}'.format(args['output']))
    e_logger.info('Finish!')


def parse_arguments():
    argument_parser = argparse.ArgumentParser(
        prog='es2exc - Elasticsearch query to Excel',
        description='Query Elasticsearch and create an excel report with the result',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
        add_help=True
    )
    argument_parser.add_argument('--version', action='version', version='%(prog)s {}'.format(__version__))

    arguments_options = argument_parser.add_subparsers()
    cli = arguments_options.add_parser('cli', help='Command line arguments')
    cli.add_argument('--host', help='host:port to make the query', default='127.0.0.1:9200')
    cli.add_argument('--index', help='index (pattern) to make the query', required=True)
    cli.add_argument('--query', help='es query to make, every aggregation will be a table', required=True)
    cli.add_argument('--output', help='output file name', default='es2exc_output.xlsx')
    cli.add_argument('--piechart', help='add a pie chart from aggregations', action='store_true')
    cli.add_argument('--barchart', help='add a bar chart from aggregations', action='store_true')

    conf = arguments_options.add_parser('conf', help='Configuration file')
    conf.add_argument('--conf', help='path to condfiguration file', default='myreport.yml')

    return vars(argument_parser.parse_args())


if __name__ == '__main__':

    args = parse_arguments()
    e_logger.info(args)

    try:
        if 'conf' in args:
            args = yaml.load(open(args['conf']))
        query = json.loads(args['query'].replace("'", '"'))
        es_client = Elasticsearch(hosts=[args['host']])
    except Exception as ex:
        e_logger.exception(ex)
        print('1')
        sys.exit(1)

    main()
